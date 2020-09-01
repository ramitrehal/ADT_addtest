import xml.etree.ElementTree as ET 
import xlsxwriter
import os.path
import openpyxl
path = 'C:\\users\\robi\\desktop\\dir\\rest\\SoapUIResults'
num_files = len([f for f in os.listdir(path)if os.path.isfile(os.path.join(path, f))])
num=num_files/2
wb = openpyxl.load_workbook("C:\\users\\robi\\desktop\\new1.xlsx") 
sheet = wb['ACTUAL'] 
#clm = sheet.max_column
rwnm = len([row for row in sheet if not all([cell.value == None for cell in row])]) + 1
#clm = len([column for column in sheet if not all([cell.value == None for cell in column])]) + 2
#print(clm)
a=0
i=1
for j in range(1,int(num+1)):
    dom = ET.parse("C:\\users\\robi\\desktop\\dir\\rest\\SoapUIResults\\response_"+str(j) +".xml")
    returnCode = dom.find('returnCode').text
    returnMessage = dom.find('returnMessage').text
    sheet.cell(row = rwnm, column= 6).value=int(returnCode)
    sheet.cell(row = rwnm, column= 7).value=returnMessage
    orderz = dom.findall('orders/order')
    a=a+(len(orderz))
    while i < a+1:
        for c in orderz:
           # matchingSample = c.find('matchingSample').text
            orderCode = c.find('orderCode').text
            orderCodeMnemonic = c.find('orderCodeMnemonic').text 
            returnCode = c.find('returnCode').text
            returnMessage = c.find('returnMessage').text
           # stabilityFlag = c.find('stabilityFlag').text
         #print (matchingSample,orderCode,orderCodeMnemonic)
           # outsheet.write(i,0,matchingSample)
            sheet.cell(row = rwnm, column= 2).value=int(orderCode)
            sheet.cell(row = rwnm, column= 3).value=orderCodeMnemonic
            sheet.cell(row = rwnm, column= 4).value=int(returnCode)
            sheet.cell(row = rwnm, column= 5).value=returnMessage
            i += 1 
            rwnm=rwnm+1
#print(returnCode)
#print(returnMessage)
# outsheet.write("A1","MATCHING SAMPLE")
# outsheet.write("B1","orderCode")
# outsheet.write("C1","orderCodeMnemonic")
# outsheet.write("A2",matchingSample)
# outsheet.write("B2",orderCode)
# outsheet.write("C2",orderCodeMnemonic)
wb.save("C:\\users\\robi\\desktop\\new1.xlsx") 
